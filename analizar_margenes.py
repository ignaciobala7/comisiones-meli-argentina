# -*- coding: utf-8 -*-
"""
ANALISIS DE MARGENES FLEXXUS vs COMISIONES MELI
Procesa el Excel de Flexxus y calcula cuanto hay que subir el margen
en Flexxus para no salir perdiendo despues de que MeLi cobre su comision.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import sys, os, time, json
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import consulta_comisiones as backend

# ─────────────────────────────────────────────
# FUNCIONES DE CALCULO CLAVE
# ─────────────────────────────────────────────

def margen_real_meli(mg_flexxus_pct, comision_pct):
    """
    Dado el margen % sobre costo que tiene Flexxus y la comision % de MeLi,
    devuelve el margen real % sobre precio de venta que queda despues de MeLi.

    Ejemplo: mg_flexxus=40%, comision=12.5% → margen real = 16.1%
    """
    if mg_flexxus_pct is None or comision_pct is None:
        return None
    factor = 1 + mg_flexxus_pct / 100        # precio = costo × factor
    neto   = factor * (1 - comision_pct / 100) # neto = precio × (1 - com%)
    return round((neto - 1) / factor * 100, 2)


def mg_flexxus_necesario(comision_pct, margen_obj_pct):
    """
    Calcula el margen % sobre costo que hay que configurar en Flexxus
    para que despues de la comision de MeLi quede margen_obj_pct% sobre precio.

    Ejemplo: comision=12.5%, objetivo=20% → necesitas 48.1% en Flexxus
    """
    denom = 1 - comision_pct / 100 - margen_obj_pct / 100
    if denom <= 0:
        return None
    return round((1 / denom - 1) * 100, 2)


def precio_venta_necesario(precio_compra_ars, comision_pct, margen_obj_pct):
    """Precio de venta en ARS que da margen_obj% despues de la comision MeLi"""
    denom = 1 - comision_pct / 100 - margen_obj_pct / 100
    if denom <= 0:
        return None
    return round(precio_compra_ars / denom, 2)


# ─────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────

def _f(bold=False, size=10, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _borde(color="D0D0D0"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _alin(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _pct(ws, row, col, valor, color_neg="FCE4D6", color_ok="E2EFDA", umbral=0):
    c = ws.cell(row, col)
    if valor is None:
        c.value = None
        return c
    c.value = float(valor) / 100
    c.number_format = "0.0%"
    c.font = _f(size=9)
    c.border = _borde("E8E8E8")
    c.fill = _fill(color_neg if float(valor) < umbral else color_ok)
    return c

def _ars(ws, row, col, valor, bold=False, color=None):
    c = ws.cell(row, col)
    if valor is None:
        c.value = None
        return c
    c.value = float(valor)
    c.number_format = "$#,##0.00"
    c.font = _f(bold=bold, size=9)
    c.border = _borde("E8E8E8")
    if color:
        c.fill = _fill(color)
    return c

def _txt(ws, row, col, valor, bold=False, size=9, wrap=False):
    c = ws.cell(row, col)
    c.value = valor if valor is not None else ""
    c.font = _f(bold=bold, size=size)
    c.border = _borde("E8E8E8")
    c.alignment = _alin(wrap=wrap)
    return c


# ─────────────────────────────────────────────
# PROCESAMIENTO PRINCIPAL
# ─────────────────────────────────────────────

IVA_DEFAULT = 21.0  # IVA general Argentina — se aplica cuando PORCENTAJEII = 0

def procesar(archivo, margen_obj=20, progress_cb=None, iva_default=IVA_DEFAULT):
    """
    Lee el Excel de Flexxus, mapea cada rubro a una categoria MeLi,
    calcula margenes reales y necesarios, y devuelve lista de dicts.
    """
    # Leer tipo de cambio
    tc = 1420.0
    try:
        xl = pd.ExcelFile(archivo)
        if "MONEDAS" in xl.sheet_names:
            df_mon = pd.read_excel(archivo, sheet_name="MONEDAS", dtype=str)
            df_mon.columns = [str(c).strip() for c in df_mon.columns]
            fila = df_mon[df_mon["CODIGOMONEDA"].str.strip().str.upper() == "DOLARES"]
            if not fila.empty:
                tc = float(str(fila["CAMBIO"].iloc[0]).replace(",", "."))
    except Exception:
        pass

    # Leer hoja ARTICULOS
    hoja = "ARTICULOS" if "ARTICULOS" in pd.ExcelFile(archivo).sheet_names else None
    df = pd.read_excel(archivo, sheet_name=hoja, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)

    # Mapeo de columnas Flexxus
    COL = {
        "sku":         "CODIGOPARTICULAR",
        "desc":        "DESCRIPCION",
        "ean":         "CODIGOBARRA",
        "costo_usd":   "PRECIOCOMPRA",
        "precio_usd":  "PRECIOVENTA5",
        "margen":      "MARGEN5",
        "iva":         "PORCENTAJEII",
        "coeficiente": "COEFICIENTE",
        "marca":       "MARCAS.DESCRIPCION",
        "rubro":       "RUBROS.DESCRIPCION",
        "superrubro":  "RUBROS.SUPERRUBROS.DESCRIPCION",
    }

    # COEFICIENTE → tasa IVA
    # 0.5 = 10.5%,  1 = 21%,  0 = 0% (exento)
    COEF_IVA = {"0.5": 10.5, "0,5": 10.5, "1": 21.0, "1.0": 21.0, "0": 0.0, "0.0": 0.0}

    # Filtrar productos sin precio o impositivos
    def es_valido(fila):
        try:
            pc = float(str(fila.get(COL["costo_usd"], "0")).replace(",", "."))
            sr = str(fila.get(COL["superrubro"], "")).upper()
            return pc > 0 and "IMPOSITIV" not in sr
        except Exception:
            return False

    filas = [r.to_dict() for _, r in df.iterrows() if es_valido(r.to_dict())]
    total = len(filas)

    # Cache: superrubro → (cat_raiz, comision_clasica, comision_premium)
    cache_rubros = {}

    resultados = []

    for i, d in enumerate(filas):
        if progress_cb:
            progress_cb(i + 1, total, str(d.get(COL["desc"], ""))[:40])

        superrubro = str(d.get(COL["superrubro"], "")).strip()
        rubro      = str(d.get(COL["rubro"], "")).strip()

        # Obtener comisiones para este rubro (con cache)
        if superrubro not in cache_rubros:
            texto_busqueda = f"{superrubro} {rubro}".strip()
            cat_id = cat_raiz = ""
            cats = backend.buscar_categoria(texto_busqueda, limite=1)
            if cats:
                cat_id   = cats[0].get("category_id", "")
                cat_raiz = backend.obtener_raiz_categoria(cat_id) if cat_id else ""
            tasas = backend.obtener_comision(cat_raiz or superrubro)
            cache_rubros[superrubro] = {
                "cat_id":   cat_id,
                "cat_raiz": cat_raiz,
                "tasas":    tasas,
            }
            time.sleep(0.12)

        info_cat = cache_rubros[superrubro]
        tasas    = info_cat["tasas"]

        # IVA desde COEFICIENTE (fuente principal)
        # 0.5 → 10.5%  |  1 → 21%  |  0 → 0%
        coef_raw = str(d.get(COL["coeficiente"], "")).strip().replace(",", ".")
        if coef_raw in COEF_IVA:
            iva = COEF_IVA[coef_raw]
        else:
            # Fallback: PORCENTAJEII, o iva_default si ambos son 0
            try:
                iva_arch = float(str(d.get(COL["iva"], "0")).replace(",", "."))
            except Exception:
                iva_arch = 0
            iva = iva_arch if iva_arch > 0 else iva_default
        factor_iva = 1 + iva / 100

        try:
            costo_usd  = float(str(d.get(COL["costo_usd"],  "0")).replace(",", "."))
            precio_usd = float(str(d.get(COL["precio_usd"], "0")).replace(",", "."))
            mg_actual  = float(str(d.get(COL["margen"],     "0")).replace(",", "."))
        except Exception:
            costo_usd = precio_usd = mg_actual = 0

        costo_ars  = round(costo_usd  * tc * factor_iva, 2)
        precio_ars = round(precio_usd * tc * factor_iva, 2)

        # Calcular por tipo de publicacion
        tipos = {}
        for tipo_nom, pct_com in tasas.items():
            mr   = margen_real_meli(mg_actual, pct_com)
            mgnec = mg_flexxus_necesario(pct_com, margen_obj)
            ajuste = round(mgnec - mg_actual, 2) if mgnec is not None else None
            pvnec  = precio_venta_necesario(costo_ars, pct_com, margen_obj)

            tipos[tipo_nom] = {
                "comision":      pct_com,
                "margen_real":   mr,
                "mg_nec":        mgnec,
                "ajuste":        ajuste,
                "precio_flexxus_ars": pvnec,
                # precio en USD sin IVA para cargar en Flexxus:
                "precio_flexxus_usd": round(pvnec / tc / factor_iva, 2) if pvnec and tc and factor_iva else None,
            }

        ean_raw = str(d.get(COL["ean"], "")).strip()
        # Limpiar EAN: solo aceptamos números puros
        # Si tiene UNA letra al principio seguida de puros dígitos → sacamos la letra
        # Si tiene letras mezcladas → descartamos
        if ean_raw in ("nan", "None", "0", ""):
            ean = ""
        elif ean_raw.isdigit():
            ean = ean_raw
        elif len(ean_raw) > 1 and ean_raw[0].isalpha() and ean_raw[1:].isdigit():
            ean = ean_raw[1:]   # e.g. "L6941264098126" → "6941264098126"
        else:
            ean = ""

        resultados.append({
            "sku":       d.get(COL["sku"], ""),
            "desc":      d.get(COL["desc"], ""),
            "ean":       ean,
            "superrubro": superrubro,
            "rubro":     rubro,
            "marca":     d.get(COL["marca"], ""),
            "cat_raiz":  info_cat["cat_raiz"],
            "costo_ars": costo_ars,
            "precio_ars": precio_ars,
            "mg_flexxus": mg_actual,
            "iva":       iva,
            "tipos":     tipos,
        })

    return resultados, tc, margen_obj


# ─────────────────────────────────────────────
# GENERAR EXCEL DE SALIDA
# ─────────────────────────────────────────────

def generar_excel(resultados, ruta_salida, margen_obj, tc):
    wb = openpyxl.Workbook()
    tipos_nom = list(resultados[0]["tipos"].keys()) if resultados else ["Clasica", "Premium"]

    # ── HOJA 1: RESUMEN POR SUPERRUBRO ──────────────────────────
    ws1 = wb.active
    ws1.title = "Por Rubro"
    ws1.freeze_panes = "B3"

    # Titulo
    n_cols_rubros = 3 + len(tipos_nom) * 4
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols_rubros)
    c = ws1.cell(1, 1)
    c.value     = f"ANALISIS POR RUBRO  —  Margen objetivo: {margen_obj}%  |  TC: ${tc:,.0f}  |  {datetime.now().strftime('%d/%m/%Y')}"
    c.font      = _f(bold=True, size=12, color="FFFFFF")
    c.fill      = _fill("1F4E79")
    c.alignment = _alin("center")
    ws1.row_dimensions[1].height = 26

    # Cabeceras
    hdrs = ["Super Rubro", "Categoria MeLi", "Cant. Prods."]
    for t in tipos_nom:
        hdrs += [
            f"Comision MeLi ({t})",
            f"Margen Flexxus prom.",
            f"Margen real en MeLi (prom.)",
            f"Margen Flexxus necesario para {margen_obj}%",
        ]
    anchos = [35, 28, 12] + [18, 18, 22, 24] * len(tipos_nom)

    col_hdrs_color = {"Clasica": "BDD7EE", "Premium": "C6EFCE"}
    # Fila de grupos
    ws1.cell(2, 1).value = ""
    ws1.cell(2, 2).value = ""
    ws1.cell(2, 3).value = ""
    ci_g = 4
    for t in tipos_nom:
        ws1.merge_cells(start_row=2, start_column=ci_g, end_row=2, end_column=ci_g + 3)
        gc = ws1.cell(2, ci_g, value=f"PUBLICACION {t.upper()}")
        gc.font = _f(bold=True, size=10, color="FFFFFF")
        gc.fill = _fill("2E75B6" if t == "Clasica" else "375623")
        gc.alignment = _alin("center")
        ci_g += 4
    ws1.row_dimensions[2].height = 18

    for ci, (h, w) in enumerate(zip(hdrs, anchos), 1):
        c = ws1.cell(3, ci, value=h)
        c.font      = _f(bold=True, size=9, color="FFFFFF")
        c.fill      = _fill("404040")
        c.alignment = _alin("center", wrap=True)
        c.border    = _borde()
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[3].height = 42

    # Agrupar por superrubro
    rubros = defaultdict(lambda: {"items": [], "cat_raiz": ""})
    for r in resultados:
        rubros[r["superrubro"]]["items"].append(r)
        if r["cat_raiz"]:
            rubros[r["superrubro"]]["cat_raiz"] = r["cat_raiz"]

    # Ordenar por margen real Clasica ascendente (peor primero)
    def sort_key(kv):
        items = kv[1]["items"]
        vals = [i["tipos"].get(tipos_nom[0], {}).get("margen_real") for i in items]
        vals = [v for v in vals if v is not None]
        return sum(vals) / len(vals) if vals else 999

    rubros_sorted = sorted(rubros.items(), key=sort_key)

    for ri, (sr, info) in enumerate(rubros_sorted, 4):
        items = info["items"]
        par = ri % 2 == 0

        ws1.cell(ri, 1).value  = sr
        ws1.cell(ri, 1).font   = _f(bold=True, size=9)
        ws1.cell(ri, 1).border = _borde()
        ws1.cell(ri, 2).value  = info["cat_raiz"]
        ws1.cell(ri, 2).font   = _f(size=9)
        ws1.cell(ri, 2).border = _borde()
        ws1.cell(ri, 3).value  = len(items)
        ws1.cell(ri, 3).font   = _f(size=9)
        ws1.cell(ri, 3).border = _borde()
        ws1.cell(ri, 3).alignment = _alin("center")

        ci = 4
        for t in tipos_nom:
            vals_com   = [i["tipos"][t]["comision"]    for i in items if t in i["tipos"]]
            vals_mg    = [i["mg_flexxus"]               for i in items]
            vals_mr    = [i["tipos"][t]["margen_real"]  for i in items if t in i["tipos"] and i["tipos"][t]["margen_real"] is not None]
            vals_mgnec = [i["tipos"][t]["mg_nec"]       for i in items if t in i["tipos"] and i["tipos"][t]["mg_nec"] is not None]

            com_prom   = sum(vals_com)   / len(vals_com)   if vals_com   else None
            mg_prom    = sum(vals_mg)    / len(vals_mg)    if vals_mg    else None
            mr_prom    = sum(vals_mr)    / len(vals_mr)    if vals_mr    else None
            mgnec_prom = sum(vals_mgnec) / len(vals_mgnec) if vals_mgnec else None
            ajuste_prom = round(mgnec_prom - mg_prom, 1) if (mgnec_prom and mg_prom) else None

            # Comision
            c = ws1.cell(ri, ci);     c.value = (com_prom or 0)/100; c.number_format = "0.0%"; c.font = _f(size=9); c.border = _borde(); c.fill = _fill("DAEEF3" if par else "EBF3FD")
            # Margen Flexxus prom
            c = ws1.cell(ri, ci+1);   c.value = (mg_prom  or 0)/100; c.number_format = "0.0%"; c.font = _f(size=9); c.border = _borde(); c.fill = _fill("FFF2CC")
            # Margen real en MeLi
            _pct(ws1, ri, ci+2, mr_prom,
                 color_neg="FCE4D6", color_ok="E2EFDA", umbral=margen_obj)
            # Margen Flexxus necesario
            if mgnec_prom is not None:
                c = ws1.cell(ri, ci+3)
                c.value = mgnec_prom / 100
                c.number_format = "0.0%"
                c.font  = _f(bold=True, size=9, color="375623" if (ajuste_prom or 0) <= 0 else "C00000")
                c.fill  = _fill("C6EFCE" if (ajuste_prom or 0) <= 0 else "FCE4D6")
                c.border = _borde()
                if ajuste_prom and ajuste_prom > 0:
                    c.comment = None
                    # Agrego el ajuste como texto adicional en la celda no, lo pongo en siguiente col
            ci += 4

    # ── HOJA 2: POR PRODUCTO (todos con ajuste requerido) ───────
    ws2 = wb.create_sheet("Ajustes por Producto")
    ws2.freeze_panes = "D4"

    hdrs2 = [
        ("SKU",               10),
        ("Descripcion",       42),
        ("Super Rubro",       28),
        ("Categoria MeLi",    22),
        ("Precio Compra ARS", 18),
        ("Precio Venta ARS",  18),
        ("Margen Flexxus %",  16),
    ]
    for t in tipos_nom:
        hdrs2 += [
            (f"Comision MeLi ({t})",     16),
            (f"Margen real ({t})",       16),
            (f"Margen Flexxus nec. ({t})",18),
            (f"↑ Subir margen ({t})",    16),
            (f"PRECIO A CARGAR ARS ({t})", 22),
            (f"PRECIO CARGAR USD ({t})", 20),
        ]

    n2 = len(hdrs2)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n2)
    c = ws2.cell(1, 1)
    c.value     = f"AJUSTES REQUERIDOS POR PRODUCTO  —  Margen objetivo: {margen_obj}%  |  TC: ${tc:,.0f}  |  {datetime.now().strftime('%d/%m/%Y')}"
    c.font      = _f(bold=True, size=12, color="FFFFFF")
    c.fill      = _fill("C00000")
    c.alignment = _alin("center")
    ws2.row_dimensions[1].height = 26

    # Grupos de cabecera
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    gc = ws2.cell(2, 1, value="DATOS FLEXXUS")
    gc.font = _f(bold=True, size=10, color="FFFFFF"); gc.fill = _fill("404040"); gc.alignment = _alin("center")
    ci_g = 8
    for t in tipos_nom:
        ws2.merge_cells(start_row=2, start_column=ci_g, end_row=2, end_column=ci_g + 5)
        gc = ws2.cell(2, ci_g, value=f"PUBLICACION {t.upper()}")
        gc.font = _f(bold=True, size=10, color="FFFFFF")
        gc.fill = _fill("2E75B6" if t == "Clasica" else "375623")
        gc.alignment = _alin("center")
        ci_g += 6
    ws2.row_dimensions[2].height = 18

    for ci, (h, w) in enumerate(hdrs2, 1):
        c = ws2.cell(3, ci, value=h)
        c.font      = _f(bold=True, size=9, color="FFFFFF")
        c.fill      = _fill("404040")
        c.alignment = _alin("center", wrap=True)
        c.border    = _borde()
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[3].height = 42

    # Filtrar solo los que necesitan ajuste en al menos un tipo
    def necesita_ajuste(r):
        for t in tipos_nom:
            aj = r["tipos"].get(t, {}).get("ajuste")
            if aj is not None and aj > 0:
                return True
        return False

    # Ordenar: primero los que necesitan más ajuste (clasica)
    prods_ajuste = [r for r in resultados if necesita_ajuste(r)]
    prods_ajuste.sort(key=lambda r: r["tipos"].get(tipos_nom[0], {}).get("margen_real") or 999)

    for ri, r in enumerate(prods_ajuste, 4):
        par = ri % 2 == 0
        fondo_base = "F0F4FA" if par else "FFFFFF"

        _txt(ws2, ri, 1, r["sku"])
        _txt(ws2, ri, 2, r["desc"], wrap=True)
        _txt(ws2, ri, 3, r["superrubro"])
        _txt(ws2, ri, 4, r["cat_raiz"])
        _ars(ws2, ri, 5, r["costo_ars"])
        _ars(ws2, ri, 6, r["precio_ars"])

        # Margen Flexxus actual
        c = ws2.cell(ri, 7)
        c.value = r["mg_flexxus"] / 100 if r["mg_flexxus"] is not None else None
        c.number_format = "0.0%"
        c.font = _f(size=9)
        c.fill = _fill("FFF2CC")
        c.border = _borde()

        ci = 8
        for t in tipos_nom:
            td = r["tipos"].get(t, {})
            com    = td.get("comision")
            mr     = td.get("margen_real")
            mgnec  = td.get("mg_nec")
            ajuste = td.get("ajuste")
            pv_ars = td.get("precio_flexxus_ars")
            pv_usd = td.get("precio_flexxus_usd")

            # Comision MeLi
            c = ws2.cell(ri, ci); c.value = (com or 0)/100; c.number_format = "0.0%"; c.font = _f(size=9); c.border = _borde(); c.fill = _fill("DAEEF3")
            # Margen real
            _pct(ws2, ri, ci+1, mr, color_neg="FCE4D6", color_ok="E2EFDA", umbral=margen_obj)
            # Margen Flexxus necesario
            c = ws2.cell(ri, ci+2)
            if mgnec:
                c.value = mgnec/100; c.number_format = "0.0%"
                c.font = _f(bold=True, size=9)
                c.fill = _fill("C6EFCE")
            c.border = _borde()
            # Ajuste necesario
            c = ws2.cell(ri, ci+3)
            if ajuste is not None:
                c.value = ajuste/100; c.number_format = "+0.0%;-0.0%;0.0%"
                c.font = _f(bold=True, size=9, color="C00000" if ajuste > 0 else "375623")
                c.fill = _fill("FCE4D6" if ajuste > 0 else "E2EFDA")
            c.border = _borde()
            # Precio a cargar en Flexxus (ARS)
            _ars(ws2, ri, ci+4, pv_ars, bold=True, color="C6EFCE")
            # Precio a cargar en Flexxus (USD sin IVA, como esta en Flexxus)
            _ars(ws2, ri, ci+5, pv_usd, bold=True, color="E2EFDA")

            ci += 6

    # ── HOJA 3: TODOS LOS PRODUCTOS ─────────────────────────────
    ws3 = wb.create_sheet("Todos los Productos")
    ws3.freeze_panes = "D4"

    # Mismas cabeceras que hoja 2
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n2)
    c = ws3.cell(1, 1)
    c.value     = f"TODOS LOS PRODUCTOS  —  Margen objetivo: {margen_obj}%  |  TC: ${tc:,.0f}  |  {datetime.now().strftime('%d/%m/%Y')}"
    c.font      = _f(bold=True, size=12, color="FFFFFF")
    c.fill      = _fill("1F4E79")
    c.alignment = _alin("center")
    ws3.row_dimensions[1].height = 26

    # Copiar estructura de cabeceras
    ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    gc = ws3.cell(2, 1, value="DATOS FLEXXUS")
    gc.font = _f(bold=True, size=10, color="FFFFFF"); gc.fill = _fill("404040"); gc.alignment = _alin("center")
    ci_g = 8
    for t in tipos_nom:
        ws3.merge_cells(start_row=2, start_column=ci_g, end_row=2, end_column=ci_g + 5)
        gc = ws3.cell(2, ci_g, value=f"PUBLICACION {t.upper()}")
        gc.font = _f(bold=True, size=10, color="FFFFFF")
        gc.fill = _fill("2E75B6" if t == "Clasica" else "375623")
        gc.alignment = _alin("center")
        ci_g += 6
    ws3.row_dimensions[2].height = 18

    for ci, (h, w) in enumerate(hdrs2, 1):
        c = ws3.cell(3, ci, value=h)
        c.font      = _f(bold=True, size=9, color="FFFFFF")
        c.fill      = _fill("404040")
        c.alignment = _alin("center", wrap=True)
        c.border    = _borde()
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.row_dimensions[3].height = 42

    for ri, r in enumerate(resultados, 4):
        par = ri % 2 == 0
        _txt(ws3, ri, 1, r["sku"])
        _txt(ws3, ri, 2, r["desc"], wrap=True)
        _txt(ws3, ri, 3, r["superrubro"])
        _txt(ws3, ri, 4, r["cat_raiz"])
        _ars(ws3, ri, 5, r["costo_ars"])
        _ars(ws3, ri, 6, r["precio_ars"])

        c = ws3.cell(ri, 7)
        c.value = r["mg_flexxus"] / 100 if r["mg_flexxus"] is not None else None
        c.number_format = "0.0%"; c.font = _f(size=9); c.fill = _fill("FFF2CC"); c.border = _borde()

        ci = 8
        for t in tipos_nom:
            td = r["tipos"].get(t, {})
            com = td.get("comision"); mr = td.get("margen_real"); mgnec = td.get("mg_nec")
            ajuste = td.get("ajuste"); pv_ars = td.get("precio_flexxus_ars"); pv_usd = td.get("precio_flexxus_usd")

            c = ws3.cell(ri, ci); c.value = (com or 0)/100; c.number_format = "0.0%"; c.font = _f(size=9); c.border = _borde(); c.fill = _fill("DAEEF3")
            _pct(ws3, ri, ci+1, mr, color_neg="FCE4D6", color_ok="E2EFDA", umbral=margen_obj)
            c = ws3.cell(ri, ci+2)
            if mgnec: c.value = mgnec/100; c.number_format = "0.0%"; c.font = _f(size=9); c.fill = _fill("C6EFCE")
            c.border = _borde()
            c = ws3.cell(ri, ci+3)
            if ajuste is not None:
                c.value = ajuste/100; c.number_format = "+0.0%;-0.0%;0.0%"
                c.font  = _f(size=9, color="C00000" if ajuste > 0 else "375623")
                c.fill  = _fill("FCE4D6" if ajuste > 0 else "E2EFDA")
            c.border = _borde()
            _ars(ws3, ri, ci+4, pv_ars, color="C6EFCE")
            _ars(ws3, ri, ci+5, pv_usd, color="E2EFDA")
            ci += 6

    # ── HOJA 4: INSTRUCTIVO ─────────────────────────────────────
    ws4 = wb.create_sheet("Como leer este reporte")
    ws4.column_dimensions["A"].width = 90

    lineas = [
        ("COMO LEER ESTE REPORTE", True, "1F4E79"),
        ("", False, None),
        ("CONCEPTO CLAVE", True, "2E75B6"),
        ("  El margen en Flexxus se calcula sobre el COSTO (markup).", False, None),
        ("  La comision de MeLi se descuenta del PRECIO DE VENTA.", False, None),
        ("  Por eso, el margen real que te queda en MeLi es MENOR que el margen de Flexxus.", False, None),
        ("", False, None),
        ("EJEMPLO PRACTICO", True, "2E75B6"),
        (f"  Producto: Tablet  |  Comision MeLi: 12.5%  |  Margen Flexxus: 40%", False, None),
        (f"  Precio compra: $100.000  →  Precio venta Flexxus: $140.000 (40% sobre costo)", False, None),
        (f"  MeLi cobra: $17.500 (12.5% sobre $140.000)", False, None),
        (f"  Vos cobras neto: $122.500", False, None),
        (f"  Tu margen REAL: ($122.500 - $100.000) / $140.000 = 16.1%  ← no es 40%!", False, None),
        ("", False, None),
        ("COLUMNAS PRINCIPALES", True, "2E75B6"),
        ("  Margen Flexxus actual %:     El % de margen que tenes configurado hoy en Flexxus (sobre costo)", False, None),
        ("  Comision MeLi:               Lo que MeLi te descuenta del precio de venta", False, None),
        ("  Margen real en MeLi:         Lo que realmente te queda a vos despues de la comision (sobre precio)", False, None),
        (f"  Margen Flexxus necesario:   El % que tenes que poner en Flexxus para llegar a {margen_obj}% neto", False, None),
        ("  ↑ Subir margen:              La diferencia — cuanto necesitas aumentar en Flexxus", False, None),
        ("  PRECIO A CARGAR ARS:         El precio de venta en pesos que tenes que poner en Flexxus", False, None),
        ("  PRECIO CARGAR USD:           El mismo precio pero en dolares SIN IVA (como lo pide Flexxus)", False, None),
        ("", False, None),
        ("COLORES", True, "2E75B6"),
        ("  ROJO en Margen real:         El margen actual es menor al objetivo", False, None),
        ("  VERDE en Margen real:        El margen actual supera el objetivo", False, None),
        ("  AMARILLO en Margen Flexxus:  El margen de Flexxus actual (referencia)", False, None),
        ("  VERDE en Precio a cargar:    El precio que tenes que ingresar en Flexxus", False, None),
        ("", False, None),
        ("HOJAS DEL REPORTE", True, "2E75B6"),
        ("  Por Rubro:             Resumen por categoria — ves de un vistazo cuales rubros estan mal", False, None),
        ("  Ajustes por Producto:  Solo los productos que necesitan subir el precio", False, None),
        ("  Todos los Productos:   Analisis completo de todos los articulos", False, None),
    ]

    for ri, (txt, bold, color) in enumerate(lineas, 1):
        c = ws4.cell(ri, 1, value=txt)
        c.font = _f(bold=bold, size=11 if bold else 10,
                    color="FFFFFF" if color else "222222")
        if color:
            c.fill = _fill(color)
            c.alignment = _alin("center")
        ws4.row_dimensions[ri].height = 22

    wb.save(ruta_salida)
    return ruta_salida


# ─────────────────────────────────────────────
# MAIN (para ejecutar directo)
# ─────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    archivo = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\ibala\Downloads\Articulos Nacho.xlsx"
    margen_obj = int(sys.argv[2]) if len(sys.argv) > 2 else 20

    print(f"\n  Procesando: {os.path.basename(archivo)}")
    print(f"  Margen objetivo: {margen_obj}%\n")

    def prog(i, total, desc):
        print(f"\r  [{i:4}/{total}] {desc:<45}", end="", flush=True)

    resultados, tc, margen_obj = procesar(archivo, margen_obj, progress_cb=prog)
    print(f"\n\n  Procesados {len(resultados)} productos.")

    ruta_sal = archivo.replace(".xlsx", f"_analisis_meli_{margen_obj}pct.xlsx")
    generar_excel(resultados, ruta_sal, margen_obj, tc)
    print(f"\n  Excel generado: {ruta_sal}")
    import subprocess
    subprocess.Popen(["start", "", ruta_sal], shell=True)
