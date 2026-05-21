# -*- coding: utf-8 -*-
"""
Crea un Excel de ejemplo con el formato esperado para importar desde Flexxus.
Ejecutar una sola vez para tener el template listo.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mis Productos"

# Columnas del template
columnas = [
    ("CODIGO",        12),
    ("DESCRIPCION",   40),
    ("RUBRO",         22),
    ("SUB_RUBRO",     22),
    ("PRECIO_VENTA",  18),
    ("COSTO_NETO",    16),
    ("STOCK",         10),
]

# Estilos header
h_font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
h_fill  = PatternFill("solid", fgColor="1F4E79")
h_align = Alignment(horizontal="center", vertical="center")
lado    = Side(style="thin", color="AAAAAA")
borde   = Border(left=lado, right=lado, top=lado, bottom=lado)

for col_idx, (nombre, ancho) in enumerate(columnas, 1):
    c = ws.cell(1, col_idx, value=nombre)
    c.font    = h_font
    c.fill    = h_fill
    c.alignment = h_align
    c.border  = borde
    ws.column_dimensions[get_column_letter(col_idx)].width = ancho

ws.row_dimensions[1].height = 28

# Filas de ejemplo
ejemplos = [
    ("PROD001", "AURICULARES BLUETOOTH JBL TUNE 510BT", "AUDIO Y SONIDO", "AURICULARES",   49999, 28000, 5),
    ("PROD002", "NOTEBOOK LENOVO IDEAPAD 3 15 I5 8GB",  "INFORMATICA",    "NOTEBOOKS",    649999, 410000, 3),
    ("PROD003", "CELULAR SAMSUNG GALAXY A15 128GB",      "TELEFONIA",      "CELULARES",    249999, 165000, 8),
    ("PROD004", "SMART TV 43 LG FULL HD",                "TV Y VIDEO",     "TELEVISORES",  459999, 290000, 2),
    ("PROD005", "ZAPATILLAS NIKE AIR MAX SC MUJER",      "INDUMENTARIA",   "CALZADO",       89999,  52000, 10),
    ("PROD006", "CAFETERA NESPRESSO ESSENZA MINI",       "ELECTRODOMEST.", "CAFETERAS",    149999,  92000, 4),
    ("PROD007", "SILLA GAMER REDRAGON SCREAM PRO",       "MUEBLES",        "SILLAS",        89999,  54000, 6),
]

d_font  = Font(name="Arial", size=10)
d_align = Alignment(vertical="center")
n_fmt   = '$#,##0.00'

for ri, fila in enumerate(ejemplos, 2):
    for ci, val in enumerate(fila, 1):
        c = ws.cell(ri, ci, value=val)
        c.font      = d_font
        c.alignment = d_align
        c.border    = borde
        if ci in (5, 6):  # Precio y costo
            c.number_format = n_fmt
        elif ci == 7:     # Stock
            c.number_format = "#,##0"
    if ri % 2 == 0:
        for ci in range(1, len(columnas) + 1):
            ws.cell(ri, ci).fill = PatternFill("solid", fgColor="EDF3F9")

ws.freeze_panes = "A2"

# Nota al pie
ws.cell(len(ejemplos) + 3, 1).value = (
    "INSTRUCCIONES: Completa este archivo con tus productos de Flexxus y ejecuta: "
    'python consulta_comisiones.py --excel template_mis_productos.xlsx'
)
ws.cell(len(ejemplos) + 3, 1).font = Font(name="Arial", italic=True, size=9, color="666666")

output = "template_mis_productos.xlsx"
wb.save(output)
print(f"\n  Template creado: {output}")
print("  Completalo con tus productos y luego ejecuta el script principal.\n")
