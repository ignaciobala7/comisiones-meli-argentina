# -*- coding: utf-8 -*-
"""
CONSULTA DE COMISIONES - MERCADO LIBRE ARGENTINA
Sistema de inteligencia de precios basado en comisiones por categoria
"""

import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import argparse
import json
import sys
import time
import os
import io
from datetime import datetime

# Forzar UTF-8 en la terminal de Windows
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stdin.reconfigure(encoding="utf-8", errors="replace")
except AttributeError:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stdin  = io.TextIOWrapper(sys.stdin.buffer,  encoding="utf-8", errors="replace")

# ============================================================
# CONFIGURACION
# ============================================================

SITE_ID  = "MLA"
BASE_URL = "https://api.mercadolibre.com"
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
TABLA_PATH = os.path.join(BASE_DIR, "comisiones_tabla.json")

_cache_cats  = {}   # texto -> lista de categorias MeLi
_cache_raiz  = {}   # category_id -> nombre de categoria raiz


def cargar_tabla():
    try:
        with open(TABLA_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"  [!] No se pudo leer {TABLA_PATH}: {e}")
        return {"_default": {"Clasica": 14.0, "Premium": 19.0}, "categorias": {}}


TABLA = cargar_tabla()
TIPOS_PUBLICACION = TABLA.get("_info", {}).get("tipos_publicacion", {
    "Clasica": "gold_special",
    "Premium": "gold_premium",
})


# ============================================================
# FUNCIONES API (solo endpoints publicos)
# ============================================================

def api_get(url, params=None, reintentos=2):
    for intento in range(reintentos):
        try:
            resp = requests.get(url, params=params, timeout=12)
            if resp.status_code == 429:
                time.sleep(2 ** intento)
                continue
            if resp.status_code == 200:
                return resp.json()
            return None
        except requests.exceptions.RequestException:
            if intento == reintentos - 1:
                return None
            time.sleep(1)
    return None


def buscar_categoria(texto, limite=8):
    clave = texto.lower().strip()
    if clave in _cache_cats:
        return _cache_cats[clave]
    data = api_get(
        f"{BASE_URL}/sites/{SITE_ID}/domain_discovery/search",
        params={"q": texto, "limit": limite},
    )
    resultado = data or []
    _cache_cats[clave] = resultado
    return resultado


def obtener_raiz_categoria(cat_id):
    """Devuelve el nombre de la categoria raiz (nivel 1) para un category_id"""
    if cat_id in _cache_raiz:
        return _cache_raiz[cat_id]
    data = api_get(f"{BASE_URL}/categories/{cat_id}")
    if data and "path_from_root" in data:
        raiz = data["path_from_root"][0]["name"] if data["path_from_root"] else data.get("name", "")
    else:
        raiz = ""
    _cache_raiz[cat_id] = raiz
    return raiz


# ============================================================
# CONSULTA DE COMISION (desde tabla local)
# ============================================================

def obtener_comision(cat_raiz):
    """
    Devuelve {tipo_nombre: pct} para una categoria raiz.
    Usa la tabla local comisiones_tabla.json.
    """
    cats = TABLA.get("categorias", {})
    # Busqueda exacta
    if cat_raiz in cats:
        return cats[cat_raiz]
    # Busqueda parcial (por si hay tildes o variaciones menores)
    cat_lower = cat_raiz.lower().strip()
    for k, v in cats.items():
        if k.lower().strip() == cat_lower:
            return v
        if cat_lower and (cat_lower in k.lower() or k.lower() in cat_lower):
            return v
    # Default
    return TABLA.get("_default", {"Clasica": 14.0, "Premium": 19.0})


def calcular_precio_sugerido(costo, pct_comision, margen_obj):
    """Precio al que publicar para obtener margen_obj% sobre el precio de venta"""
    denom = 1 - pct_comision / 100 - margen_obj / 100
    if denom <= 0:
        return None
    return round(costo / denom, 2)


# ============================================================
# MODO INTERACTIVO (consulta individual)
# ============================================================

def sep(char="─", n=60):
    print("  " + char * n)


def modo_interactivo():
    print()
    sep("═")
    print("  CONSULTA INDIVIDUAL DE COMISIONES")
    sep("═")

    # 1. Buscar categoria
    while True:
        query = input(
            "\n  Que producto queres consultar?\n"
            "  (ej: 'celular samsung', 'notebook', 'zapatillas'): "
        ).strip()
        if not query:
            continue

        print("  Buscando en Mercado Libre...", end="", flush=True)
        cats = buscar_categoria(query)
        print("\r" + " " * 50 + "\r", end="")

        if not cats:
            print("  No se encontraron categorias. Proba con otro termino.")
            continue

        print(f"\n  Categorias encontradas para '{query}':")
        sep()
        for i, cat in enumerate(cats, 1):
            dominio = cat.get("domain_name", "")
            nombre  = cat.get("category_name", "")
            cat_id  = cat.get("category_id", "")
            display = f"{dominio} > {nombre}" if dominio and nombre and dominio != nombre else (dominio or nombre)
            print(f"  [{i:2}]  {display:<46} {cat_id}")
        sep()

        try:
            op = input("  Elegi el numero (Enter para buscar de nuevo): ").strip()
            if not op:
                continue
            idx = int(op) - 1
            if 0 <= idx < len(cats):
                cat_sel = cats[idx]
                break
        except ValueError:
            pass
        print("  Opcion invalida.")

    cat_id     = cat_sel.get("category_id")
    cat_nombre = cat_sel.get("domain_name") or cat_sel.get("category_name", cat_id)

    # Obtener categoria raiz para buscar en la tabla
    cat_raiz = obtener_raiz_categoria(cat_id)
    tasas    = obtener_comision(cat_raiz)

    # 2. Precio de venta
    while True:
        try:
            precio = float(
                input("\n  A que precio lo vas a publicar? $: ")
                .strip().replace(",", ".").replace("$", "").replace(" ", "")
            )
            if precio > 0:
                break
        except ValueError:
            pass
        print("  Ingresa un numero valido.")

    # 3. Costo (opcional)
    costo = None
    try:
        s = input("\n  Cual es tu costo? (Enter para omitir) $: ").strip()
        if s:
            costo = float(s.replace(",", ".").replace("$", "").replace(" ", ""))
    except ValueError:
        pass

    # 4. Mostrar resultados
    print()
    sep("═")
    print(f"  PRODUCTO:          {query.upper()}")
    print(f"  CATEGORIA MeLi:    {cat_nombre}  ({cat_id})")
    print(f"  CATEGORIA RAIZ:    {cat_raiz or 'N/A'}")
    print(f"  PRECIO DE VENTA:   ${precio:>12,.2f}")
    if costo:
        print(f"  COSTO:             ${costo:>12,.2f}")
    sep("═")

    for tipo_nombre, pct in tasas.items():
        fee  = round(precio * pct / 100, 2)
        neto = round(precio - fee, 2)

        sep()
        print(f"  PUBLICACION {tipo_nombre.upper()}")
        sep()
        print(f"  Comision MeLi:    {pct:5.1f}%   ->   ${fee:>12,.2f}")
        print(f"  Lo que cobras:                   ${neto:>12,.2f}")

        if costo and costo > 0:
            ganancia = neto - costo
            margen   = ganancia / precio * 100
            roi      = ganancia / costo * 100
            signo    = "GANAS" if ganancia >= 0 else "PERDES"
            print(f"  Tu ganancia:      {margen:5.1f}%   ->   ${ganancia:>12,.2f}   ({signo})")
            print(f"  ROI sobre costo:  {roi:5.1f}%")
            print()
            print(f"  Precios sugeridos para distintos margenes objetivo:")
            for m in [10, 15, 20, 25, 30]:
                ps = calcular_precio_sugerido(costo, pct, m)
                if ps:
                    print(f"    {m:2}% margen  ->  Publicar a  ${ps:>10,.2f}")
                else:
                    print(f"    {m:2}% margen  ->  Imposible (comision + margen >= 100%)")

    sep("═")
    info = TABLA.get("_info", {})
    print(f"  Tasas segun tabla local  |  Ultima actualizacion: {info.get('ultima_actualizacion', 'N/A')}")
    print(f"  Verificar en: {info.get('fuente', 'https://www.mercadolibre.com.ar')}")


# ============================================================
# MODO EXCEL (FLEXXUS)
# ============================================================

PALABRAS_CLAVE = {
    "codigo":      ["codigoparticular", "cod", "codigo", "código", "sku", "artículo", "articulo", "codart", "id_art"],
    "descripcion": ["desc", "nombre", "denominac", "detalle", "product", "item"],
    "rubro":       ["superrubros", "rubro", "categoria", "categoría", "departamento", "familia", "grupo", "linea"],
    "subrubro":    ["rubros.descripcion", "subrub", "subcategor", "subgrupo", "sublinea", "subfamil"],
    "precio":      ["precioventa", "precio", "pvp", "p.v.", "lista", "venta", "sale_price", "p_vta"],
    "costo":       ["preciocompra", "costo", "cost", "compra", "neto_comp", "p_comp"],
    "stock":       ["stock", "cantidad", "qty", "existencia", "disponib"],
    "iva":         ["porcentajeii", "iva", "alicuota", "tasa_iva"],
    "margen_orig": ["margen5", "margen", "margin"],
}


def detectar_columnas(cols):
    lower_map = {str(c): str(c).lower() for c in cols}
    encontrado = {}
    for tipo, palabras in PALABRAS_CLAVE.items():
        for orig, low in lower_map.items():
            for p in palabras:
                if p in low:
                    encontrado[tipo] = orig
                    break
            if tipo in encontrado:
                break
    return encontrado


def limpiar_numero(valor):
    if valor is None or str(valor).strip() in ("", "nan", "None", "-"):
        return None
    try:
        s = str(valor).strip().replace("$", "").replace(" ", "")
        if "," in s and "." in s:
            if s.rindex(",") > s.rindex("."):
                s = s.replace(".", "").replace(",", ".")  # 1.234,56 -> 1234.56
            else:
                s = s.replace(",", "")                    # 1,234.56 -> 1234.56
        elif "," in s:
            s = s.replace(",", ".")                       # 1234,56 -> 1234.56
        v = float(s)
        return v if v >= 0 else None
    except (ValueError, AttributeError):
        return None


def procesar_excel(archivo, margen_obj):
    print(f"\n  Leyendo: {os.path.basename(archivo)}")

    # Detectar si hay hoja ARTICULOS (formato Flexxus nativo)
    hoja = None
    tc = 1.0
    es_usd = False
    try:
        xl = pd.ExcelFile(archivo)
        if "ARTICULOS" in xl.sheet_names:
            hoja = "ARTICULOS"
        if "MONEDAS" in xl.sheet_names:
            df_mon = pd.read_excel(archivo, sheet_name="MONEDAS", dtype=str)
            df_mon.columns = [str(c).strip() for c in df_mon.columns]
            if "CAMBIO" in df_mon.columns and "CODIGOMONEDA" in df_mon.columns:
                fila_dolar = df_mon[df_mon["CODIGOMONEDA"].str.strip().str.upper() == "DOLARES"]
                if not fila_dolar.empty:
                    tc = float(str(fila_dolar["CAMBIO"].iloc[0]).replace(",", "."))
                    es_usd = True
    except Exception as e:
        print(f"  [!] No se pudo leer tipo de cambio: {e}")

    try:
        df = pd.read_excel(archivo, sheet_name=hoja, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how="all").reset_index(drop=True)
    except Exception as e:
        print(f"  [ERROR] No se pudo leer el archivo: {e}")
        return None

    print(f"  Filas:    {len(df)}")
    print(f"  Columnas: {list(df.columns)}")
    if es_usd:
        print(f"  Moneda:   USD  |  Tipo de cambio: ${tc:,.0f}")
    print()

    col_map = detectar_columnas(df.columns)
    print("  Columnas detectadas automaticamente:")
    for tipo, col in col_map.items():
        print(f"    {tipo:<14} ->  '{col}'")

    # Pedir columnas criticas que no se detectaron
    for campo, desc in [("rubro", "rubro/categoria"), ("precio", "precio de venta")]:
        if campo not in col_map:
            print(f"\n  ATENCION: No se detecto la columna de {desc}.")
            print("  Columnas disponibles:")
            for i, c in enumerate(df.columns, 1):
                print(f"    [{i}] {c}")
            try:
                idx = int(input(f"  Numero de la columna '{desc}': ")) - 1
                if 0 <= idx < len(df.columns):
                    col_map[campo] = df.columns[idx]
            except ValueError:
                pass

    # Procesar productos
    # Cache: texto_rubro -> (cat_id, cat_nombre, cat_raiz, tasas)
    mapa_cats = {}
    resultados = []
    total = len(df)

    print(f"\n  Procesando {total} productos...")

    for idx, fila in df.iterrows():
        d = fila.to_dict()
        dato = {c: ("" if str(d.get(c, "")) == "nan" else d.get(c, "")) for c in df.columns}

        precio = limpiar_numero(d.get(col_map.get("precio", "")))
        costo  = limpiar_numero(d.get(col_map.get("costo", "")))

        # Conversión USD → ARS + IVA (formato Flexxus nativo)
        if es_usd and tc > 1:
            iva_pct = limpiar_numero(d.get(col_map.get("iva", ""), 0)) or 0
            factor = tc * (1 + iva_pct / 100)
            if precio: precio = round(precio * factor, 2)
            if costo:  costo  = round(costo  * factor, 2)

        # Guardar margen original de Flexxus
        dato["__margen_flexxus"] = limpiar_numero(d.get(col_map.get("margen_orig", ""), None))

        # Filtrar artículos sin precio (servicios, impositivos, etc.)
        if (precio is None or precio == 0) and (costo is None or costo == 0):
            continue

        # Armar texto de busqueda
        texto = ""
        if "rubro" in col_map:
            texto = str(d.get(col_map["rubro"], "")).strip()
            if "subrubro" in col_map:
                sub = str(d.get(col_map["subrubro"], "")).strip()
                if sub and sub != "nan":
                    texto = f"{texto} {sub}".strip()
        if not texto or texto in ("nan", "None"):
            if "descripcion" in col_map:
                texto = str(d.get(col_map["descripcion"], ""))[:50].strip()

        cat_id, cat_nombre, cat_raiz, tasas = "", "No encontrada", "", TABLA.get("_default", {})

        if texto and texto not in ("nan", "None"):
            if texto in mapa_cats:
                cat_id, cat_nombre, cat_raiz, tasas = mapa_cats[texto]
            else:
                print(f"\r  [{idx+1:4}/{total}] Buscando: '{texto[:38]}'...{' '*10}", end="", flush=True)
                cats = buscar_categoria(texto, limite=1)
                if cats:
                    cat_id     = cats[0].get("category_id", "")
                    cat_nombre = cats[0].get("domain_name") or cats[0].get("category_name", "")
                    cat_raiz   = obtener_raiz_categoria(cat_id) if cat_id else ""
                    tasas      = obtener_comision(cat_raiz)
                mapa_cats[texto] = (cat_id, cat_nombre, cat_raiz, tasas)
                time.sleep(0.15)

        dato["__cat_id"]    = cat_id
        dato["__cat_nombre"] = cat_nombre
        dato["__cat_raiz"]  = cat_raiz
        dato["__precio"]    = precio
        dato["__costo"]     = costo

        for tipo_nombre, pct in tasas.items():
            p = tipo_nombre.lower()
            dato[f"__pct_{p}"] = pct

            if precio and precio > 0:
                fee  = round(precio * pct / 100, 2)
                neto = round(precio - fee, 2)
                dato[f"__fee_{p}"]  = fee
                dato[f"__neto_{p}"] = neto
                if costo and costo > 0:
                    margen = (neto - costo) / precio * 100
                    dato[f"__margen_{p}"] = round(margen, 2)
                    dato[f"__sug_{p}"]    = calcular_precio_sugerido(costo, pct, margen_obj)
                else:
                    dato[f"__margen_{p}"] = None
                    dato[f"__sug_{p}"]    = None
            else:
                dato[f"__fee_{p}"]    = None
                dato[f"__neto_{p}"]   = None
                dato[f"__margen_{p}"] = None
                dato[f"__sug_{p}"]    = None

        resultados.append(dato)
        desc = str(d.get(col_map.get("descripcion", ""), ""))[:35]
        print(f"\r  [{idx+1:4}/{total}] OK  {desc:<35}", end="", flush=True)

    print(f"\n\n  Procesados {total} productos.")
    return resultados, col_map, list(df.columns)


# ============================================================
# GENERAR EXCEL DE SALIDA
# ============================================================

C_AZUL_OSC = "1F4E79"
C_AZUL_MED = "2E75B6"
C_VERDE_OSC = "375623"
C_GRIS      = "404040"
C_AZUL_CL   = "DAEEF3"
C_VERDE_CL  = "E2EFDA"
C_ROJO_CL   = "FCE4D6"
C_AMARILLO  = "FFF2CC"
C_ALTERNO   = "F0F4FA"


def _f(bold=False, size=10, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _borde(color="D0D0D0"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _alin(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _set_pct(cell, valor, margen_obj=None):
    """Asigna un porcentaje (valor en %) al cell con formato adecuado y color si es margen"""
    if valor is None:
        cell.value = None
        return
    try:
        v = float(valor)
        cell.value = v / 100
        cell.number_format = "0.0%"
        if margen_obj is not None:
            if v < 0:
                cell.fill = _fill(C_ROJO_CL)
            elif v >= margen_obj:
                cell.fill = _fill(C_VERDE_CL)
            else:
                cell.fill = _fill(C_AMARILLO)
    except (ValueError, TypeError):
        cell.value = valor


def _set_ars(cell, valor, fondo=None):
    if valor is None:
        cell.value = None
        return
    try:
        cell.value = float(valor)
        cell.number_format = "$#,##0.00"
        if fondo:
            cell.fill = _fill(fondo)
    except (ValueError, TypeError):
        cell.value = valor


def generar_excel_salida(resultados, col_map, cols_orig, ruta, margen_obj):
    wb = openpyxl.Workbook()

    tipos_nombres = list(TIPOS_PUBLICACION.keys())

    # ─────────────────────────────────────────────────────────
    # HOJA 1: ACCION REQUERIDA (productos con margen bajo)
    # ─────────────────────────────────────────────────────────
    ws_acc = wb.active
    ws_acc.title = "Accion Requerida"
    ws_acc.freeze_panes = "A4"

    # Filtrar y ordenar por margen Clasica ascendente (peor primero)
    p_clasica = tipos_nombres[0].lower() if tipos_nombres else "clasica"
    def _margen_sort(d):
        v = d.get(f"__margen_{p_clasica}")
        return float(v) if v is not None else 999

    ajuste = [d for d in resultados
              if any(d.get(f"__margen_{t.lower()}") is not None and
                     float(d.get(f"__margen_{t.lower()}")) < margen_obj
                     for t in tipos_nombres)]
    ajuste.sort(key=_margen_sort)

    col_cod  = col_map.get("codigo", "")
    col_desc = col_map.get("descripcion", "")
    col_prec = col_map.get("precio", "")
    col_cost = col_map.get("costo", "")

    # Cabeceras hoja accion
    acc_headers = [
        ("Codigo",           12, "id"),
        ("Descripcion",      40, "id"),
        ("Categoria MeLi",   28, "id"),
        ("Precio Actual ($)", 18, "id"),
        ("Costo ($)",        14, "id"),
    ]
    for t in tipos_nombres:
        p = t.lower()
        acc_headers += [
            (f"Margen Actual {t}",        14, t),
            (f"Brecha al {margen_obj}% {t}", 14, t),
            (f"PRECIO A CARGAR ({t})",    20, t),
            (f"Diferencia ($) {t}",       16, t),
        ]

    N_acc = len(acc_headers)

    # Titulo
    ws_acc.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_acc)
    c = ws_acc.cell(1, 1)
    c.value     = f"PRODUCTOS QUE NECESITAN AJUSTE DE PRECIO  —  Margen objetivo: {margen_obj}%   |   {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font      = _f(bold=True, size=12, color="FFFFFF")
    c.fill      = _fill("C00000")
    c.alignment = _alin("center")
    ws_acc.row_dimensions[1].height = 28

    # Grupo de cabeceras
    ci_g = 1
    col_acc_color = {"Clasica": C_AZUL_MED, "Premium": C_VERDE_OSC}
    ws_acc.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    gc = ws_acc.cell(2, 1, value="DATOS DEL PRODUCTO")
    gc.font = _f(bold=True, size=10, color="FFFFFF")
    gc.fill = _fill(C_GRIS)
    gc.alignment = _alin("center")
    ci_g = 6
    for t in tipos_nombres:
        ws_acc.merge_cells(start_row=2, start_column=ci_g, end_row=2, end_column=ci_g + 3)
        gc = ws_acc.cell(2, ci_g, value=f"PUBLICACION {t.upper()}")
        gc.font = _f(bold=True, size=10, color="FFFFFF")
        gc.fill = _fill(col_acc_color.get(t, C_AZUL_MED))
        gc.alignment = _alin("center")
        ci_g += 4
    ws_acc.row_dimensions[2].height = 20

    fill_acc = {"id": "D0D0D0", "Clasica": C_AZUL_CL, "Premium": "C6EFCE"}
    for ci, (h, w, grp) in enumerate(acc_headers, 1):
        c = ws_acc.cell(3, ci, value=h)
        c.font      = _f(bold=True, size=9)
        c.alignment = _alin("center", wrap=True)
        c.border    = _borde()
        c.fill      = _fill(fill_acc.get(grp, "DDDDDD"))
        ws_acc.column_dimensions[get_column_letter(ci)].width = w
    ws_acc.row_dimensions[3].height = 42

    for ri, dato in enumerate(ajuste, 4):
        precio_act = dato.get("__precio")
        costo_val  = dato.get("__costo")

        fila_vals = [
            dato.get(col_cod, ""),
            dato.get(col_desc, ""),
            dato.get("__cat_nombre", ""),
            precio_act,
            costo_val,
        ]
        for t in tipos_nombres:
            p = t.lower()
            mg  = dato.get(f"__margen_{p}")
            sug = dato.get(f"__sug_{p}")
            brecha = (float(mg) - margen_obj) if mg is not None else None
            dif    = (float(sug) - float(precio_act)) if (sug and precio_act) else None
            fila_vals += [mg, brecha, sug, dif]

        for ci, v in enumerate(fila_vals, 1):
            c = ws_acc.cell(ri, ci)
            c.font   = _f(size=9)
            c.border = _borde("E8E8E8")
            c.alignment = _alin()

            grp = acc_headers[ci - 1][2]
            col_name = acc_headers[ci - 1][0]

            if v is None or str(v) in ("nan", "None"):
                c.value = None
                continue

            # Margen actual → colorear
            if "Margen Actual" in col_name:
                try:
                    vf = float(v)
                    c.value = vf / 100
                    c.number_format = "0.0%"
                    c.fill = _fill(C_ROJO_CL if vf < 0 else C_AMARILLO)
                except (ValueError, TypeError):
                    c.value = v

            # Brecha → rojo si negativa
            elif "Brecha" in col_name:
                try:
                    vf = float(v)
                    c.value = vf / 100
                    c.number_format = "0.0%"
                    c.fill = _fill(C_ROJO_CL if vf < 0 else C_AMARILLO)
                except (ValueError, TypeError):
                    c.value = v

            # Precio a cargar → destacado verde
            elif "PRECIO A CARGAR" in col_name:
                try:
                    c.value = float(v)
                    c.number_format = "$#,##0.00"
                    c.fill = _fill("C6EFCE")
                    c.font = _f(bold=True, size=10, color="375623")
                except (ValueError, TypeError):
                    c.value = v

            # Diferencia
            elif "Diferencia" in col_name:
                try:
                    vf = float(v)
                    c.value = vf
                    c.number_format = "$#,##0.00"
                    c.fill = _fill("FCE4D6" if vf > 0 else "E2EFDA")
                except (ValueError, TypeError):
                    c.value = v

            # Precio / Costo
            elif col_name in ("Precio Actual ($)", "Costo ($)"):
                try:
                    c.value = float(v)
                    c.number_format = "$#,##0.00"
                except (ValueError, TypeError):
                    c.value = v

            else:
                c.value = v

    # Nota si no hay productos para ajustar
    if not ajuste:
        ws_acc.merge_cells(start_row=4, start_column=1, end_row=4, end_column=N_acc)
        c = ws_acc.cell(4, 1)
        c.value     = f"Todos los productos tienen margen >= {margen_obj}%. No se requieren ajustes."
        c.font      = _f(bold=True, size=11, color="375623")
        c.fill      = _fill(C_VERDE_CL)
        c.alignment = _alin("center")

    # ─────────────────────────────────────────────────────────
    # HOJA 2: ANALISIS POR PRODUCTO
    # ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Analisis Completo")
    ws.freeze_panes = "C4"

    # Definir columnas de salida
    prioridad = ["codigo", "descripcion", "rubro", "subrubro", "precio", "costo", "stock"]
    labels_map = {
        "codigo":      "Codigo",
        "descripcion": "Descripcion",
        "rubro":       "Rubro",
        "subrubro":    "Sub-Rubro",
        "precio":      "Precio Lista ($)",
        "costo":       "Costo ($)",
        "stock":       "Stock",
    }
    cols_flex = []
    usadas = set()
    for t in prioridad:
        if t in col_map:
            cols_flex.append((labels_map[t], col_map[t], "flex"))
            usadas.add(col_map[t])
    for c in cols_orig:
        if c not in usadas:
            cols_flex.append((c, c, "flex"))

    # Incluir margen Flexxus si existe en los datos
    tiene_mg_flexxus = any(d.get("__margen_flexxus") is not None for d in resultados)
    cols_meli = [
        ("Margen Flexxus %",  "__margen_flexxus", "meli") if tiene_mg_flexxus else None,
        ("Categoria MeLi",    "__cat_nombre",     "meli"),
        ("Categoria Raiz",    "__cat_raiz",        "meli"),
        ("ID Categoria",      "__cat_id",          "meli"),
    ]
    cols_meli = [c for c in cols_meli if c is not None]

    cols_tipos = []
    for t in tipos_nombres:
        p = t.lower()
        cols_tipos += [
            (f"Comision %",                   f"__pct_{p}",    t),
            (f"Comision $",                   f"__fee_{p}",    t),
            (f"Neto cobras $",                f"__neto_{p}",   t),
            (f"Tu Margen %",                  f"__margen_{p}", t),
            (f"Precio Flexxus ({margen_obj}%)", f"__sug_{p}",  t),
        ]

    todas = cols_flex + cols_meli + cols_tipos
    N = len(todas)

    FILA_T = 1
    FILA_G = 2
    FILA_H = 3
    FILA_D = 4

    # Titulo
    ws.merge_cells(start_row=FILA_T, start_column=1, end_row=FILA_T, end_column=N)
    c = ws.cell(FILA_T, 1)
    c.value = (
        f"ANALISIS DE COMISIONES  —  MERCADO LIBRE ARGENTINA"
        f"   |   {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        f"   |   Margen objetivo: {margen_obj}%"
    )
    c.font      = _f(bold=True, size=12, color="FFFFFF")
    c.fill      = _fill(C_AZUL_OSC)
    c.alignment = _alin("center")
    ws.row_dimensions[FILA_T].height = 28

    # Grupos
    nf = len(cols_flex)
    nm = len(cols_meli)
    col_color = {"Clasica": C_AZUL_MED, "Premium": C_VERDE_OSC}
    grupos = [
        ("DATOS (Flexxus)", 1, nf, C_GRIS),
        ("MERCADO LIBRE",   nf + 1, nf + nm, C_AZUL_MED),
    ]
    ci = nf + nm + 1
    for t in tipos_nombres:
        nt = sum(1 for _, _, g in cols_tipos if g == t)
        grupos.append((f"PUBLICACION {t.upper()}", ci, ci + nt - 1, col_color.get(t, C_AZUL_MED)))
        ci += nt

    for label, ci_g, cf_g, color in grupos:
        if ci_g < cf_g:
            ws.merge_cells(start_row=FILA_G, start_column=ci_g, end_row=FILA_G, end_column=cf_g)
        gc = ws.cell(FILA_G, ci_g, value=label)
        gc.font      = _f(bold=True, size=10, color="FFFFFF")
        gc.fill      = _fill(color)
        gc.alignment = _alin("center")
    ws.row_dimensions[FILA_G].height = 20

    # Headers
    anchos = {
        "Codigo": 12, "Descripcion": 38, "Rubro": 20, "Sub-Rubro": 18,
        "Precio Lista ($)": 16, "Costo ($)": 14, "Stock": 10,
        "Categoria MeLi": 30, "Categoria Raiz": 26, "ID Categoria": 14,
        "Comision %": 12, "Comision $": 13, "Neto cobras $": 15, "Tu Margen %": 13,
        "Margen Flexxus %": 14,
    }
    fill_grupo = {
        "flex": "D0D0D0",
        "meli": "BDD7EE",
        "Clasica": C_AZUL_CL,
        "Premium": "C6EFCE",
    }
    for ci, (header, key, grupo) in enumerate(todas, 1):
        c = ws.cell(FILA_H, ci, value=header)
        c.font      = _f(bold=True, size=9)
        c.alignment = _alin("center", wrap=True)
        c.border    = _borde()
        c.fill      = _fill(fill_grupo.get(grupo, "DDDDDD"))
        ancho = anchos.get(header, 14)
        if "Sug" in header:
            ancho = 18
        ws.column_dimensions[get_column_letter(ci)].width = ancho
    ws.row_dimensions[FILA_H].height = 36

    # Datos
    for ri, dato in enumerate(resultados, FILA_D):
        par = ri % 2 == 0
        for ci, (header, key, grupo) in enumerate(todas, 1):
            c = ws.cell(ri, ci)
            val = dato.get(key)
            if str(val) in ("nan", "None", "none"):
                val = None
            c.font      = _f(size=9)
            c.alignment = _alin()
            c.border    = _borde("E8E8E8")

            if key.startswith("__pct_"):
                c.fill = _fill("EBF3FD" if grupo == "Clasica" else "EBF5EB")
                _set_pct(c, val)

            elif key.startswith("__margen_"):
                _set_pct(c, val, margen_obj=margen_obj)

            elif key.startswith("__fee_") or key.startswith("__neto_"):
                fondo = "EBF3FD" if grupo == "Clasica" else "EBF5EB"
                _set_ars(c, val, fondo=fondo)

            elif key.startswith("__sug_"):
                fondo = "D6E8FB" if grupo == "Clasica" else "D5EED5"
                _set_ars(c, val, fondo=fondo)

            elif key == "__margen_flexxus":
                _set_pct(c, val)
                c.fill = _fill("FFF2CC")  # amarillo: margen de origen Flexxus

            elif key == col_map.get("precio") or key == col_map.get("costo"):
                _set_ars(c, val)
                if par:
                    c.fill = _fill(C_ALTERNO)

            elif key == col_map.get("stock"):
                if val is not None:
                    try:
                        c.value = float(val)
                        c.number_format = "#,##0"
                    except (ValueError, TypeError):
                        c.value = val
                if par:
                    c.fill = _fill(C_ALTERNO)

            else:
                c.value = val or ""
                if par:
                    c.fill = _fill(C_ALTERNO)

    # Nota al pie
    fila_nota = FILA_D + len(resultados) + 1
    ws.merge_cells(start_row=fila_nota, start_column=1, end_row=fila_nota, end_column=N)
    info = TABLA.get("_info", {})
    ws.cell(fila_nota, 1).value = (
        f"Comisiones segun tabla local  |  "
        f"Ultima actualizacion: {info.get('ultima_actualizacion', 'N/A')}  |  "
        f"Verificar en: {info.get('fuente', 'https://www.mercadolibre.com.ar')}"
    )
    ws.cell(fila_nota, 1).font      = _f(size=8, color="666666", bold=False)
    ws.cell(fila_nota, 1).alignment = _alin("center")
    ws.cell(fila_nota, 1).fill      = _fill("F5F5F5")

    # ─────────────────────────────────────────────────────────
    # HOJA 2: RESUMEN POR CATEGORIA
    # ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen por Categoria")
    ws2.freeze_panes = "B3"

    stats = defaultdict(lambda: {
        "count": 0,
        **{f"pct_{t.lower()}": []    for t in tipos_nombres},
        **{f"margen_{t.lower()}": [] for t in tipos_nombres},
    })
    for dato in resultados:
        cat = dato.get("__cat_raiz") or dato.get("__cat_nombre") or "Sin categoria"
        stats[cat]["count"] += 1
        for t in tipos_nombres:
            p = t.lower()
            for campo_src, campo_dst in [(f"__pct_{p}", f"pct_{p}"), (f"__margen_{p}", f"margen_{p}")]:
                v = dato.get(campo_src)
                if v is not None:
                    try:
                        stats[cat][campo_dst].append(float(v))
                    except (ValueError, TypeError):
                        pass

    headers2 = ["Categoria MeLi", "Cant."]
    for t in tipos_nombres:
        headers2 += [f"% Com. {t}", f"Margen {t} (prom)"]

    n2 = len(headers2)
    ws2.merge_cells(f"A1:{get_column_letter(n2)}1")
    c = ws2.cell(1, 1, value="RESUMEN DE COMISIONES POR CATEGORIA")
    c.font = _f(bold=True, size=12, color="FFFFFF")
    c.fill = _fill(C_AZUL_OSC)
    c.alignment = _alin("center")
    ws2.row_dimensions[1].height = 28

    for ci, h in enumerate(headers2, 1):
        c = ws2.cell(2, ci, value=h)
        c.font      = _f(bold=True, size=10, color="FFFFFF")
        c.fill      = _fill(C_GRIS)
        c.alignment = _alin("center", wrap=True)
        c.border    = _borde()
        ws2.column_dimensions[get_column_letter(ci)].width = 30 if ci == 1 else 20
    ws2.row_dimensions[2].height = 30

    for ri, (cat_n, s) in enumerate(sorted(stats.items()), 3):
        fila = [cat_n, s["count"]]
        for t in tipos_nombres:
            p = t.lower()
            lp = s.get(f"pct_{p}", [])
            lm = s.get(f"margen_{p}", [])
            fila.append(sum(lp) / len(lp)   if lp else None)
            fila.append(sum(lm) / len(lm) if lm else None)

        for ci, v in enumerate(fila, 1):
            c = ws2.cell(ri, ci, value=v)
            c.font      = _f(size=10)
            c.border    = _borde()
            c.alignment = _alin()
            if ci > 2 and v is not None:
                try:
                    vf = float(v)
                    c.value = vf / 100
                    c.number_format = "0.0%"
                    if "Margen" in headers2[ci - 1]:
                        if vf < 0:
                            c.fill = _fill(C_ROJO_CL)
                        elif vf >= margen_obj:
                            c.fill = _fill(C_VERDE_CL)
                        else:
                            c.fill = _fill(C_AMARILLO)
                except (ValueError, TypeError):
                    pass
        if ri % 2 == 0:
            for ci in range(1, n2 + 1):
                cell = ws2.cell(ri, ci)
                if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = _fill(C_ALTERNO)

    # ─────────────────────────────────────────────────────────
    # HOJA 3: INSTRUCCIONES
    # ─────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Instrucciones")
    ws3.column_dimensions["A"].width = 88

    info = TABLA.get("_info", {})
    filas_inst = [
        ("COMO INTERPRETAR ESTE REPORTE", True, C_AZUL_OSC),
        ("", False, None),
        ("COLUMNAS PRINCIPALES", True, C_AZUL_MED),
        ("  Categoria Raiz:       La gran categoria de MeLi (ej: Celulares y Telefonos, Computacion)", False, None),
        ("  Comision %:           Porcentaje que cobra MeLi sobre tu precio de venta", False, None),
        ("  Comision $:           Monto en pesos que se lleva MeLi", False, None),
        ("  Neto cobras $:        Lo que recibis vos = Precio - Comision", False, None),
        ("  Tu Margen %:          Ganancia sobre precio = (Neto - Costo) / Precio", False, None),
        (f"  Precio Flexxus ({margen_obj}%): El precio exacto que tenes que cargar en Flexxus para tener {margen_obj}% de margen", False, None),
        ("", False, None),
        ("HOJA 'ACCION REQUERIDA'", True, "C00000"),
        ("  Es la primera hoja. Muestra solo los productos con margen menor al objetivo.", False, None),
        ("  Ordenados de peor a mejor margen. La columna verde 'PRECIO A CARGAR' es el", False, None),
        ("  valor exacto que tenes que ingresar en Flexxus para cada producto.", False, None),
        ("  La columna 'Diferencia' muestra cuanto tenes que subir el precio actual.", False, None),
        ("", False, None),
        ("COLORES EN 'TU MARGEN %'", True, C_AZUL_MED),
        ("  VERDE:   Margen >= objetivo  ->  BIEN, publica a ese precio", False, None),
        ("  AMARILLO: Margen positivo pero menor al objetivo  ->  Evaluar si conviene", False, None),
        ("  ROJO:    Margen negativo  ->  A ese precio PERDES plata", False, None),
        ("", False, None),
        ("SOBRE LAS COMISIONES", True, C_AZUL_MED),
        (f"  Tasas obtenidas de tabla local  |  Actualizada: {info.get('ultima_actualizacion', 'N/A')}", False, None),
        (f"  Fuente oficial: {info.get('fuente', 'https://www.mercadolibre.com.ar')}", False, None),
        ("  Para actualizar tasas, editar el archivo 'comisiones_tabla.json' en la misma carpeta.", False, None),
        ("  Las tasas van aprox. de 10% (Alimentos) a 17% (Relojes/Joyeria) en Clasica.", False, None),
        ("  Premium tiene comisiones aproximadamente 5% mas altas que Clasica.", False, None),
        ("", False, None),
        ("COMO USAR EL SCRIPT", True, C_AZUL_MED),
        ("  Modo Excel (para tu stock de Flexxus):", False, None),
        ('    python consulta_comisiones.py --excel "mi_stock.xlsx" --margen 20', False, None),
        ("", False, None),
        ("  Modo consulta individual (sin Excel):", False, None),
        ("    python consulta_comisiones.py", False, None),
        ("", False, None),
        ("  O simplemente abre 'iniciar.bat' con doble click.", False, None),
    ]

    for ri, (texto, bold, color) in enumerate(filas_inst, 1):
        c = ws3.cell(ri, 1, value=texto)
        c.font = _f(bold=bold, size=11 if bold else 10, color="FFFFFF" if color else "222222")
        if color:
            c.fill      = _fill(color)
            c.alignment = _alin("center")
        ws3.row_dimensions[ri].height = 22

    # Guardar
    try:
        wb.save(ruta)
        return True
    except PermissionError:
        alt = ruta.replace(".xlsx", f"_v{int(time.time())}.xlsx")
        wb.save(alt)
        print(f"\n  [!] Archivo en uso. Guardado como: {os.path.basename(alt)}")
        return True
    except Exception as e:
        print(f"\n  [ERROR al guardar] {e}")
        return False


# ============================================================
# MAIN
# ============================================================

def banner():
    print("\n" + "═" * 64)
    print("  COMISIONES MERCADO LIBRE ARGENTINA")
    print("  Sistema de Inteligencia de Precios")
    print(f"  {datetime.now().strftime('%d/%m/%Y  %H:%M')}")
    info = TABLA.get("_info", {})
    print(f"  Tabla de tasas: actualizada {info.get('ultima_actualizacion', 'N/A')}")
    print("═" * 64)


def main():
    parser = argparse.ArgumentParser(description="Comisiones MeLi Argentina")
    parser.add_argument("--excel",  "-e", metavar="ARCHIVO",
                        help="Excel de Flexxus a procesar")
    parser.add_argument("--salida", "-s", metavar="ARCHIVO",
                        default="comisiones_meli.xlsx",
                        help="Nombre del Excel de salida (default: comisiones_meli.xlsx)")
    parser.add_argument("--margen", "-m", type=float, default=20,
                        help="Margen objetivo en %% para precio sugerido (default: 20)")
    args = parser.parse_args()

    banner()

    # Verificar conexion (endpoint publico de categorias)
    print("\n  Verificando conexion con Mercado Libre...")
    test = api_get(f"{BASE_URL}/sites/{SITE_ID}/domain_discovery/search", params={"q": "celular", "limit": 1})
    if test is not None:
        print("  Conexion OK  (API de categorias disponible)")
    else:
        print("  Sin conexion  —  Funciones de busqueda limitadas")

    if args.excel:
        archivo = os.path.abspath(args.excel)
        if not os.path.exists(archivo):
            print(f"\n  ERROR: No se encontro '{archivo}'")
            sys.exit(1)

        ruta_salida = os.path.join(os.path.dirname(archivo), args.salida)
        print(f"\n  Margen objetivo:   {args.margen}%")
        print(f"  Archivo de salida: {os.path.basename(ruta_salida)}")

        resultado = procesar_excel(archivo, args.margen)
        if resultado:
            resultados, col_map, cols_orig = resultado
            print("\n  Generando Excel de salida...")
            if generar_excel_salida(resultados, col_map, cols_orig, ruta_salida, args.margen):
                print(f"\n  {'═'*50}")
                print(f"  LISTO: {ruta_salida}")
                print(f"  {'═'*50}")
                print("  3 hojas generadas:")
                print("    1. 'Analisis Comisiones'   — Detalle por producto")
                print("    2. 'Resumen por Categoria' — Totales y promedios")
                print("    3. 'Instrucciones'         — Como interpretar")
    else:
        print("\n  Modo CONSULTA INDIVIDUAL.")
        print("  TIP: Para procesar tu Excel de Flexxus ejecuta:")
        print('       python consulta_comisiones.py --excel "mi_stock.xlsx"\n')
        while True:
            modo_interactivo()
            resp = input("\n  Consultar otro producto? [S/N]: ").strip().lower()
            if resp not in ("s", "si", "sí", "y"):
                break

    print("\n  Hasta luego!\n")


if __name__ == "__main__":
    main()
